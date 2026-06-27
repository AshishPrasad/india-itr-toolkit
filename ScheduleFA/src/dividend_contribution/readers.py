"""Read dividend and stock-ledger workbooks into plain Python structures."""
import openpyxl

from .parsing import parse_amount, parse_date
from .headers import find_header_row, is_blank_row, is_total_row


# --------------------------------------------------------------------------- #
# Column specifications (explicit, default_exact, aliases)
# --------------------------------------------------------------------------- #
def _dividend_specs(cols):
    cols = cols or {}
    return {
        "date": (cols.get("date"), "Date", ["date"]),
        "usd": (
            cols.get("usd"),
            "Total Dividend (USD)",
            ["total dividend (usd)", "dividend (usd)", "gross dividend (usd)",
             "amount (usd)", "usd amount", "dividend usd"],
        ),
        "rate": (
            cols.get("rate"),
            "USD -> INR (TT Buy SBI)",
            ["usd -> inr", "usd->inr", "tt buy", "exchange rate",
             "usd to inr", "fx rate", "conversion rate"],
        ),
    }


def _stock_specs(cols):
    cols = cols or {}
    return {
        "date": (cols.get("date"), "Date", ["date"]),
        "type": (
            cols.get("type"),
            "Buy/Sell",
            ["buy/sell", "buy / sell", "type", "transaction", "action", "side"],
        ),
        "lot": (
            cols.get("lot"),
            "Lot number",
            ["lot number", "lot no", "lot id", "lot"],
        ),
        "qty": (
            cols.get("qty"),
            "Quantity",
            ["quantity", "qty", "shares", "units", "no of shares"],
        ),
    }


class ReadError(Exception):
    """Raised when a workbook cannot be parsed."""


# --------------------------------------------------------------------------- #
# Worksheet extraction (testable without files)
# --------------------------------------------------------------------------- #
def extract_dividends(ws, cols=None):
    """Extract ``[(date, inr, usd, rate), ...]`` from a dividend worksheet."""
    specs = _dividend_specs(cols)
    hdr, mapping = find_header_row(ws, specs)
    if hdr is None:
        raise ReadError(
            "Could not locate dividend headers (need date, USD amount and "
            "USD->INR rate). Found fields: %s. "
            "Override with --div-date-col / --div-usd-col / --div-rate-col."
            % sorted(mapping)
        )
    rows = []
    started = False
    for r in range(hdr + 1, ws.max_row + 1):
        if is_blank_row(ws, r):
            if started:
                break
            continue
        if is_total_row(ws, r, mapping["date"]):
            break
        d = parse_date(ws.cell(r, mapping["date"]).value)
        usd = parse_amount(ws.cell(r, mapping["usd"]).value)
        rate = parse_amount(ws.cell(r, mapping["rate"]).value)
        if d is None or usd is None or rate is None:
            continue
        rows.append((d, usd * rate, usd, rate))
        started = True
    if not rows:
        raise ReadError("No dividend rows found.")
    return rows


def extract_transactions(ws, cols=None):
    """Extract ``([(date, type, lot, qty), ...], lot_order)`` from a ledger."""
    specs = _stock_specs(cols)
    hdr, mapping = find_header_row(ws, specs)
    if hdr is None:
        raise ReadError(
            "Could not locate stock-ledger headers (need Date, Buy/Sell, "
            "Lot number and Quantity). Found fields: %s. "
            "Override with --stk-date-col / --stk-type-col / --stk-lot-col / "
            "--stk-qty-col." % sorted(mapping)
        )
    txns = []
    lot_order = []
    started = False
    for r in range(hdr + 1, ws.max_row + 1):
        if is_blank_row(ws, r):
            if started:
                break
            continue
        if is_total_row(ws, r, mapping["date"]):
            break
        d = parse_date(ws.cell(r, mapping["date"]).value)
        raw_type = ws.cell(r, mapping["type"]).value
        lot = ws.cell(r, mapping["lot"]).value
        qty = parse_amount(ws.cell(r, mapping["qty"]).value)
        if d is None or raw_type is None or lot is None or qty is None:
            continue
        ttype = _normalise_type(raw_type, r)
        lot_key = _normalise_lot(lot)
        if lot_key not in lot_order:
            lot_order.append(lot_key)
        txns.append((d, ttype, lot_key, qty))
        started = True
    if not txns:
        raise ReadError("No stock transactions found.")
    return txns, lot_order


def _normalise_type(raw_type, row):
    t = str(raw_type).strip().lower()
    if t.startswith("b"):
        return "buy"
    if t.startswith("s"):
        return "sell"
    raise ReadError("Row %d has an unknown Buy/Sell value: %r." % (row, raw_type))


def _normalise_lot(lot):
    if isinstance(lot, float) and lot.is_integer():
        lot = int(lot)
    return str(lot).strip()


# --------------------------------------------------------------------------- #
# File-level readers
# --------------------------------------------------------------------------- #
def _open_sheet(path, sheet):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise ReadError("Could not open '%s': %s" % (path, exc))
    if sheet:
        if sheet not in wb.sheetnames:
            raise ReadError(
                "Sheet '%s' not found in '%s'. Available: %s"
                % (sheet, path, wb.sheetnames)
            )
        return wb[sheet]
    return wb[wb.sheetnames[0]]


def read_dividends(path, sheet=None, cols=None):
    """Read a dividend workbook into ``[(date, inr, usd, rate), ...]``."""
    ws = _open_sheet(path, sheet)
    try:
        return extract_dividends(ws, cols)
    except ReadError as exc:
        raise ReadError("%s (file: %s, sheet: %s)" % (exc, path, ws.title))


def read_transactions(path, sheet=None, cols=None):
    """Read a stock ledger into ``([(date, type, lot, qty), ...], lot_order)``."""
    ws = _open_sheet(path, sheet)
    try:
        return extract_transactions(ws, cols)
    except ReadError as exc:
        raise ReadError("%s (file: %s, sheet: %s)" % (exc, path, ws.title))
