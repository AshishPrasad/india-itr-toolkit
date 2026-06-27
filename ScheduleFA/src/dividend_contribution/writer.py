"""Write the per-lot dividend contribution result to an Excel workbook."""
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_MONEY_FMT = "\u20b9#,##0.00"
_QTY_FMT = "#,##0.0000"

# Leading characters a spreadsheet may interpret as the start of a formula.
# Includes tab (\t) and carriage return (\r), which can be used to smuggle a
# formula past naive filters (per OWASP CSV-injection guidance).
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_text(value):
    """Neutralise spreadsheet formula injection in user-supplied text cells.

    If the text starts with a character a spreadsheet treats as a formula
    (``=``, ``+``, ``-``, ``@``) or a leading tab/carriage return, prefix it
    with a single quote so it is rendered literally. Returns non-text values
    unchanged.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def write_output(path, result):
    """Write a :class:`~dividend_contribution.computation.Result` to ``path``.

    The sheet has one row per lot (lot number, net held quantity, dividend
    contribution in INR) followed by a bold grand-total row.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dividend Contribution"
    bold = Font(bold=True)

    headers = ["Lot number", "Net Held Qty", "Dividend Contribution (INR)"]
    for col, text in enumerate(headers, start=1):
        ws.cell(1, col, text).font = bold

    for row, lot in enumerate(result.lot_order, start=2):
        ws.cell(row, 1, _sanitize_text(lot))
        qty_cell = ws.cell(row, 2, round(result.net_held.get(lot, 0.0), 4))
        qty_cell.number_format = _QTY_FMT
        val_cell = ws.cell(row, 3, result.contrib_rounded.get(lot, 0.0))
        val_cell.number_format = _MONEY_FMT

    total_row = len(result.lot_order) + 2
    label = ws.cell(total_row, 1, "Total")
    label.font = bold
    qty_total = ws.cell(
        total_row, 2, round(sum(result.net_held.values()), 4)
    )
    qty_total.number_format = _QTY_FMT
    qty_total.font = bold
    val_total = ws.cell(total_row, 3, round(result.allocated, 2))
    val_total.number_format = _MONEY_FMT
    val_total.font = bold

    for col, width in ((1, 16), (2, 16), (3, 28)):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    wb.save(path)
    return path
