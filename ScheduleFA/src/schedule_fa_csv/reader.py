"""Read the per-lot Schedule FA input workbook into structured records.

The Excel-parsing helpers (amount/date parsing, column resolution, blank/total
row detection) are shared with the ``dividend_contribution`` package to avoid
duplication.
"""
import datetime
from dataclasses import dataclass

import openpyxl

from dividend_contribution.parsing import norm, parse_amount, parse_date
from dividend_contribution.headers import (
    is_blank_row,
    is_total_row,
    resolve_column,
)

_MAX_SCAN = 50


class ReadError(Exception):
    """Raised when the input workbook cannot be parsed."""


@dataclass
class Holding:
    """One foreign-equity holding (a single lot) for Schedule FA Table A3."""

    lot: str
    country_region: str
    country_code: str
    entity_name: str
    address: str
    zip_code: str
    nature: str
    date_acquired: datetime.date
    initial_value: float
    peak_value: float
    closing_value: float
    gross_paid: float
    gross_proceeds: float


# Column specs: field -> (explicit_override, default_exact, [aliases]).
# Aliases are substring-matched against normalised header text.
def _required_specs(cols):
    cols = cols or {}
    return {
        "date": (
            cols.get("date"), "Date",
            ["date of acquiring", "date acquired", "acquisition date", "date"],
        ),
        "lot": (
            cols.get("lot"), "Lot number",
            ["lot number", "lot no", "lot id", "lot"],
        ),
        "initial_value": (
            cols.get("initial_value"), "Initial value",
            ["initial value of the investment", "initial value",
             "cost basis", "cost value", "cost"],
        ),
        "peak_value": (
            cols.get("peak_value"), "Peak value in INR",
            ["peak value of investment", "peak value in inr",
             "peak value (inr)", "peak value"],
        ),
        "closing_value": (
            cols.get("closing_value"), "Closing value in INR",
            ["closing balance", "closing value in inr",
             "closing value (inr)", "closing value"],
        ),
        "gross_paid": (
            cols.get("gross_paid"), "Dividend Contribution",
            ["total gross amount paid", "dividend contribution (inr)",
             "dividend contribution", "gross amount paid"],
        ),
        "country_region": (
            cols.get("country_region"), "Country/Region name",
            ["country/region name", "country/region", "country region"],
        ),
        "country_code": (
            cols.get("country_code"), "Country Name and Code",
            ["country name and code", "country name & code",
             "country and code", "country code"],
        ),
        "entity_name": (
            cols.get("entity_name"), "Name of entity",
            ["name of entity", "entity name"],
        ),
        "address": (
            cols.get("address"), "Address of entity",
            ["address of entity", "address"],
        ),
        "zip_code": (
            cols.get("zip_code"), "ZIP Code",
            ["zip code", "zip", "postal code", "pin code", "pincode"],
        ),
        "nature": (
            cols.get("nature"), "Nature of entity",
            ["nature of entity", "nature"],
        ),
    }


def _optional_specs(cols):
    cols = cols or {}
    return {
        "gross_proceeds": (
            cols.get("gross_proceeds"), "Sale proceeds",
            ["total gross proceeds from sale", "sale proceeds in inr",
             "sale proceeds (inr)", "sale proceeds", "gross proceeds",
             "proceeds", "sale or redemption"],
        ),
    }


def _header_map(ws, row):
    mapping = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row, c).value
        if val is not None:
            mapping[norm(val)] = c
    return mapping


def _check_collisions(mapping):
    """Raise if two logical fields resolved to the same worksheet column."""
    by_col = {}
    for field, col in mapping.items():
        by_col.setdefault(col, []).append(field)
    clashes = {col: fields for col, fields in by_col.items() if len(fields) > 1}
    if clashes:
        detail = "; ".join(
            "column %d <- %s" % (col, ", ".join(sorted(fields)))
            for col, fields in sorted(clashes.items())
        )
        raise ReadError(
            "Ambiguous header mapping: %s. Rename the columns or set explicit "
            "names with the matching --*-col flags." % detail
        )


def _detect_header(ws, required, optional):
    """Find the first row resolving all required fields; add optional ones."""
    best_hit, best_missing = -1, list(required)
    for r in range(1, min(ws.max_row, _MAX_SCAN) + 1):
        header_map = _header_map(ws, r)
        mapping, missing = {}, []
        for field, spec in required.items():
            col = resolve_column(header_map, spec)
            if col is None:
                missing.append(field)
            else:
                mapping[field] = col
        if not missing:
            for field, spec in optional.items():
                col = resolve_column(header_map, spec)
                if col is not None:
                    mapping[field] = col
            _check_collisions(mapping)
            return r, mapping
        if len(required) - len(missing) > best_hit:
            best_hit = len(required) - len(missing)
            best_missing = missing
    raise ReadError(
        "Could not locate required Schedule FA input headers. Missing: %s. "
        "Use the matching --*-col flags to set custom column names."
        % sorted(best_missing)
    )


def _text(ws, row, mapping, field):
    col = mapping.get(field)
    if col is None:
        return ""
    val = ws.cell(row, col).value
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        val = int(val)  # avoid "98052.0" for numeric ZIP codes
    return str(val).strip()


def _num(ws, row, mapping, field, default=0.0):
    col = mapping.get(field)
    if col is None:
        return default
    val = parse_amount(ws.cell(row, col).value)
    return default if val is None else val


def extract_holdings(ws, cols=None):
    """Extract ``[Holding, ...]`` from a worksheet."""
    required = _required_specs(cols)
    optional = _optional_specs(cols)
    hdr, mapping = _detect_header(ws, required, optional)

    holdings = []
    started = False
    date_col = mapping["date"]
    for r in range(hdr + 1, ws.max_row + 1):
        if is_blank_row(ws, r):
            if started:
                break
            continue
        if is_total_row(ws, r, date_col):
            break
        d = parse_date(ws.cell(r, date_col).value)
        if d is None:
            continue  # skip rows without a valid acquisition date
        holdings.append(
            Holding(
                lot=_text(ws, r, mapping, "lot"),
                country_region=_text(ws, r, mapping, "country_region"),
                country_code=_text(ws, r, mapping, "country_code"),
                entity_name=_text(ws, r, mapping, "entity_name"),
                address=_text(ws, r, mapping, "address"),
                zip_code=_text(ws, r, mapping, "zip_code"),
                nature=_text(ws, r, mapping, "nature"),
                date_acquired=d,
                initial_value=_num(ws, r, mapping, "initial_value"),
                peak_value=_num(ws, r, mapping, "peak_value"),
                closing_value=_num(ws, r, mapping, "closing_value"),
                gross_paid=_num(ws, r, mapping, "gross_paid"),
                gross_proceeds=_num(ws, r, mapping, "gross_proceeds", 0.0),
            )
        )
        started = True
    if not holdings:
        raise ReadError("No data rows found in the input workbook.")
    return holdings


def _open_sheet(path, sheet):
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface a friendly message
        raise ReadError("Could not open '%s': %s" % (path, exc))
    if sheet:
        if sheet not in wb.sheetnames:
            raise ReadError(
                "Sheet '%s' not found in '%s'. Available: %s"
                % (sheet, path, wb.sheetnames)
            )
        return wb[sheet]
    return wb[wb.sheetnames[0]]


def read_holdings(path, sheet=None, cols=None):
    """Read the Schedule FA input workbook into ``[Holding, ...]``."""
    ws = _open_sheet(path, sheet)
    try:
        return extract_holdings(ws, cols)
    except ReadError as exc:
        raise ReadError("%s (file: %s, sheet: %s)" % (exc, path, ws.title))
