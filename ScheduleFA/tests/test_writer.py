"""Tests for the output writer, including formula-injection hardening."""
import datetime

import openpyxl
import pytest

from dividend_contribution.writer import _sanitize_text, write_output
from dividend_contribution.computation import compute

D = datetime.date


@pytest.mark.parametrize(
    "value, expected",
    [
        ("=1+1", "'=1+1"),
        ("+danger", "'+danger"),
        ("-cmd", "'-cmd"),
        ("@SUM", "'@SUM"),
        ("\tcmd", "'\tcmd"),     # leading tab
        ("\rcmd", "'\rcmd"),     # leading carriage return
        ("123", "123"),          # normal lot id untouched
        ("LOT-A", "LOT-A"),       # hyphen not at start untouched
        (42, 42),                # non-text untouched
    ],
)
def test_sanitize_text(value, expected):
    assert _sanitize_text(value) == expected


def test_write_output_neutralises_formula_lot_labels(tmp_path):
    # A lot label beginning with '=' must be written as inert text ("'=...").
    dividends = [(D(2025, 3, 13), 1000.0, 10.0, 100.0)]
    txns = [
        (D(2025, 1, 1), "buy", "=2+5", 10.0),
        (D(2025, 1, 1), "buy", "1", 10.0),
    ]
    lot_order = ["=2+5", "1"]
    result = compute(dividends, txns, lot_order)

    out = tmp_path / "out.xlsx"
    write_output(str(out), result)

    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws.cell(2, 1).value == "'=2+5"   # sanitised
    assert ws.cell(3, 1).value == "1"        # untouched
