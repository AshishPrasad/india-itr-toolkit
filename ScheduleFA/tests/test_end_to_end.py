"""End-to-end tests driving the CLI and reading the output workbook."""
import openpyxl

from dividend_contribution.cli import main


def _read_output(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = {}
    total = None
    for r in range(2, ws.max_row + 1):
        lot = ws.cell(r, 1).value
        held = ws.cell(r, 2).value
        contrib = ws.cell(r, 3).value
        if lot == "Total":
            total = contrib
        elif lot is not None:
            rows[str(lot)] = {"held": held, "contrib": contrib}
    return rows, total


def test_cli_end_to_end(dividends_path, stocks_path, tmp_path, capsys):
    out = tmp_path / "result.xlsx"
    rc = main([
        "--dividends", dividends_path,
        "--stocks", stocks_path,
        "--output", str(out),
    ])
    assert rc == 0
    assert out.exists()

    rows, total = _read_output(str(out))
    assert rows["1"]["contrib"] == 18416.67
    assert rows["2"]["contrib"] == 16583.33
    assert rows["1"]["held"] == 5.0
    assert rows["2"]["held"] == 10.0
    assert total == 35000.0

    captured = capsys.readouterr().out
    assert "Allocated to lots" in captured
    assert "35,000.00" in captured


def test_cli_with_renamed_headers_via_overrides(
    dividends_path, stocks_renamed_path, tmp_path
):
    out = tmp_path / "result2.xlsx"
    rc = main([
        "--dividends", dividends_path,
        "--stocks", stocks_renamed_path,
        "--output", str(out),
        "--stk-date-col", "Txn Date",
        "--stk-type-col", "Action",
        "--stk-lot-col", "Lot",
        "--stk-qty-col", "Shares",
    ])
    assert rc == 0
    rows, total = _read_output(str(out))
    assert total == 35000.0
    assert rows["1"]["contrib"] == 18416.67


def test_cli_missing_input_returns_error_code(
    dividends_path, tmp_path, capsys
):
    out = tmp_path / "result3.xlsx"
    rc = main([
        "--dividends", dividends_path,
        "--stocks", str(tmp_path / "nope.xlsx"),
        "--output", str(out),
    ])
    assert rc == 2
    assert "ERROR" in capsys.readouterr().err
