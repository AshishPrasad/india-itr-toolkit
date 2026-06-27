"""Tests that the committed template header files stay in sync with the code.

These guard against the templates under ``ScheduleFA/templates/`` drifting out of
sync with the tool (e.g. a column added to the output ``HEADERS`` but the
template not regenerated).
"""
import csv
import importlib.util
import os

import openpyxl
import pytest

from schedule_fa_csv.reader import (
    ReadError,
    _optional_specs,
    _required_specs,
    read_holdings,
)
from schedule_fa_csv.writer import HEADERS as OUTPUT_HEADERS

_TEMPLATES = os.path.join(os.path.dirname(__file__), "..", "templates")
_INPUT_XLSX = os.path.join(_TEMPLATES, "fa_input_template.xlsx")
_OUTPUT_CSV = os.path.join(_TEMPLATES, "fa_output_headers.csv")


def _load_generator():
    path = os.path.join(_TEMPLATES, "generate_templates.py")
    spec = importlib.util.spec_from_file_location("generate_templates", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _input_template_headers():
    ws = openpyxl.load_workbook(_INPUT_XLSX).active
    return [c.value for c in ws[1] if c.value is not None]


def _output_template_headers():
    with open(_OUTPUT_CSV, newline="", encoding="utf-8-sig") as fh:
        return next(csv.reader(fh))


def test_output_header_template_matches_writer_headers():
    # The committed CSV header reference must equal the tool's actual output.
    assert _output_template_headers() == OUTPUT_HEADERS


def test_input_template_matches_generator_headers():
    gen = _load_generator()
    assert _input_template_headers() == gen.INPUT_HEADERS


def test_input_template_contains_all_reader_default_headers():
    # Every default (exact) header the reader looks for must be in the template,
    # so template users always hit the safe exact-match path.
    headers = set(_input_template_headers())
    defaults = [spec[1] for spec in _required_specs(None).values()]
    defaults += [spec[1] for spec in _optional_specs(None).values()]
    missing = [d for d in defaults if d not in headers]
    assert missing == []


def test_input_template_is_header_only():
    # The committed template must ship headers only (no real/sample data rows).
    with pytest.raises(ReadError):
        read_holdings(_INPUT_XLSX)


def test_generated_templates_match_committed(tmp_path):
    # Regenerating into a fresh dir must reproduce the committed header content.
    gen = _load_generator()
    gen.generate(str(tmp_path))
    with open(tmp_path / "fa_output_headers.csv", newline="",
              encoding="utf-8-sig") as fh:
        regenerated_out = next(csv.reader(fh))
    regenerated_in = [
        c.value for c in openpyxl.load_workbook(
            str(tmp_path / "fa_input_template.xlsx")
        ).active[1] if c.value is not None
    ]
    assert regenerated_out == _output_template_headers()
    assert regenerated_in == _input_template_headers()
