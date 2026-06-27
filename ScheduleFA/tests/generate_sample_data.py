"""Generate the sample Excel workbooks used by the test suite.

Run directly to (re)create the files under the gitignored ``tests/.tmp_data``::

    python tests/generate_sample_data.py

The numbers are chosen so per-lot contributions are easy to verify by hand:

    Dividends (INR = USD * rate)
        2025-02-15  USD 100 * 80 =  8000
        2025-05-20  USD 100 * 85 =  8500
        2025-08-20  USD 100 * 90 =  9000
        2025-11-20  USD 100 * 95 =  9500   (total 35000)

    Ledger
        2025-01-01  Buy  lot 1  qty 10
        2025-05-01  Buy  lot 2  qty 10
        2025-08-01  Sell lot 1  qty 5

    Expected contribution
        lot 1 = 8000 + 4250 + 3000 + 3166.67 = 18416.67
        lot 2 =    0 + 4250 + 6000 + 6333.33 = 16583.33
"""
import datetime
import os

import openpyxl

# Sample workbooks are generated artifacts, not committed; they live in a
# gitignored temp folder alongside the tests.
DATA_DIR = os.path.join(os.path.dirname(__file__), ".tmp_data")

DIVIDENDS = [
    (datetime.datetime(2025, 2, 15), 100.0, 80.0),
    (datetime.datetime(2025, 5, 20), 100.0, 85.0),
    (datetime.datetime(2025, 8, 20), 100.0, 90.0),
    (datetime.datetime(2025, 11, 20), 100.0, 95.0),
]

LEDGER = [
    (datetime.datetime(2025, 1, 1), "Buy", 1, 10.0),
    (datetime.datetime(2025, 5, 1), "Buy", 2, 10.0),
    (datetime.datetime(2025, 8, 1), "Sell", 1, 5.0),
]

# Synthetic per-lot input for the Schedule FA CSV generator. All entity details
# are fictional. Columns:
#   date, lot, qty, initial, peak, closing, dividend, proceeds,
#   country/region, country name+code, entity, address, zip, nature
FA_HOLDINGS = [
    (datetime.datetime(2023, 5, 10), 1, 10.0, 100000.0, 150000.0, 130000.0,
     5000.0, 0.0, "United States of America",
     "United States of America - 2", "Example Corp",
     "1 Example Way, Example City", "00000", "Listed company"),
    (datetime.datetime(2024, 8, 15), 2, 5.0, 60000.0, 90000.0, 80000.0,
     2000.0, 75000.0, "United States of America",
     "United States of America - 2", "Example Corp",
     "1 Example Way, Example City", "00000", "Listed company"),
]

FA_HEADERS = [
    "Date", "Lot number", "Quantity", "Initial value", "Peak value in INR",
    "Closing value in INR", "Dividend Contribution", "Sale proceeds",
    "Country/Region name", "Country Name and Code", "Name of entity",
    "Address of entity", "ZIP Code", "Nature of entity",
]


def _money(value):
    return "\u20b9{:,.2f}".format(value)


def make_dividends(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dividends"
    # A leading title row + blank row exercise header auto-detection.
    ws["A1"] = "Dividend Schedule CY2025"
    headers = ["Date", "Total Dividend (USD)", "USD -> INR (TT Buy SBI)",
               "Total Dividend (INR)"]
    for col, text in enumerate(headers, start=1):
        ws.cell(3, col, text)
    row = 4
    total_inr = 0.0
    for d, usd, rate in DIVIDENDS:
        ws.cell(row, 1, d).number_format = "yyyy-mm-dd"
        ws.cell(row, 2, usd)
        ws.cell(row, 3, _money(rate))        # rate stored as currency text
        ws.cell(row, 4, _money(usd * rate))  # extra column, ignored by the tool
        total_inr += usd * rate
        row += 1
    ws.cell(row, 1, "Total")
    ws.cell(row, 4, _money(total_inr))
    wb.save(path)


def make_ledger(path, headers=None):
    headers = headers or ["Date", "Buy/Sell", "Lot number", "Quantity"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    for col, text in enumerate(headers, start=1):
        ws.cell(1, col, text)
    row = 2
    for d, ttype, lot, qty in LEDGER:
        ws.cell(row, 1, d).number_format = "yyyy-mm-dd"
        ws.cell(row, 2, ttype)
        ws.cell(row, 3, lot)
        ws.cell(row, 4, qty)
        row += 1
    wb.save(path)


def make_fa_input(path, headers=None):
    headers = headers or FA_HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holdings"
    for col, text in enumerate(headers, start=1):
        ws.cell(1, col, text)
    row = 2
    for rec in FA_HOLDINGS:
        d = rec[0]
        ws.cell(row, 1, d).number_format = "yyyy-mm-dd"
        # ZIP code is written as text to preserve any leading zeros.
        for col, value in enumerate(rec[1:], start=2):
            ws.cell(row, col, value)
        row += 1
    wb.save(path)


def generate(data_dir=DATA_DIR):
    os.makedirs(data_dir, exist_ok=True)
    make_dividends(os.path.join(data_dir, "dividends_sample.xlsx"))
    make_ledger(os.path.join(data_dir, "stocks_sample.xlsx"))
    # Renamed headers exercise alias matching / column overrides.
    make_ledger(
        os.path.join(data_dir, "stocks_renamed.xlsx"),
        headers=["Txn Date", "Action", "Lot", "Shares"],
    )
    make_fa_input(os.path.join(data_dir, "fa_input_sample.xlsx"))
    return data_dir


if __name__ == "__main__":
    out = generate()
    print("Sample data written to", out)
