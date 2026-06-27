"""Regenerate the header-only Schedule FA templates.

Run from anywhere::

    python ScheduleFA/templates/generate_templates.py

Produces, alongside this script:
    fa_input_template.xlsx   -- empty input sheet with the expected headers
    fa_output_headers.csv    -- the exact output column headers (one row)

The output headers are imported from ``schedule_fa_csv.writer`` so the template
always matches what the tool emits. These files contain headers only (no data)
and are safe to commit. Copy the input template OUT of the repo before entering
real financial data.
"""
import csv
import os
import sys

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Make the package importable when run directly (templates/ -> ../src).
_SRC = os.path.join(os.path.dirname(__file__), "..", "src")
sys.path.insert(0, os.path.abspath(_SRC))

from schedule_fa_csv.writer import HEADERS as OUTPUT_HEADERS  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))

# Expected input columns (one row per lot/holding). "Quantity" is optional and
# ignored by the tool but kept for convenience; "Sale proceeds" is optional.
INPUT_HEADERS = [
    "Date", "Lot number", "Quantity", "Initial value", "Peak value in INR",
    "Closing value in INR", "Dividend Contribution", "Sale proceeds",
    "Country/Region name", "Country Name and Code", "Name of entity",
    "Address of entity", "ZIP Code", "Nature of entity",
]
_INPUT_WIDTHS = [12, 10, 9, 14, 16, 18, 20, 14, 22, 26, 30, 30, 10, 18]


def make_input_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Holdings"
    for col, text in enumerate(INPUT_HEADERS, start=1):
        ws.cell(1, col, text).font = Font(bold=True)
    # Keep ZIP Code as text so leading zeros are preserved when typed.
    zip_col = INPUT_HEADERS.index("ZIP Code") + 1
    ws.cell(2, zip_col).number_format = "@"
    for col, width in enumerate(_INPUT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def make_output_headers(path):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        csv.writer(fh).writerow(OUTPUT_HEADERS)


def generate(out_dir=HERE):
    make_input_template(os.path.join(out_dir, "fa_input_template.xlsx"))
    make_output_headers(os.path.join(out_dir, "fa_output_headers.csv"))
    return out_dir


if __name__ == "__main__":
    out = generate()
    print("Templates written to", out)
